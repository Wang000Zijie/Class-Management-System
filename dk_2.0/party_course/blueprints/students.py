from datetime import datetime
from io import BytesIO

from flask import Blueprint, flash, redirect, render_template, request, send_file, url_for
from flask_login import login_required
from openpyxl import Workbook, load_workbook

from models import Assignment, Attendance, CourseSession, ExamRecord, FinalScore, Semester, Student, VolunteerRecord, db


students_bp = Blueprint("students_bp", __name__)


def get_current_semester():
	return Semester.query.filter_by(status="active").order_by(Semester.id.desc()).first()


def _ensure_student_assignments(student: Student, semester: Semester) -> None:
	sessions = CourseSession.query.filter_by(semester_id=semester.id).all()
	existing = {
		(item.type, item.session_id)
		for item in Assignment.query.filter_by(student_id=student.id, semester_id=semester.id).all()
	}

	for session in sessions:
		key = ("课堂笔记", session.id)
		if key not in existing:
			db.session.add(
				Assignment(
					student_id=student.id,
					semester_id=semester.id,
					session_id=session.id,
					type="课堂笔记",
					status="未提交",
				)
			)

	for tp in ["个人心得", "小组讨论纪要"]:
		key = (tp, None)
		if key not in existing:
			db.session.add(
				Assignment(
					student_id=student.id,
					semester_id=semester.id,
					session_id=None,
					type=tp,
					status="未提交",
				)
			)


def _require_semester_redirect():
	semester = get_current_semester()
	if not semester:
		flash("请先在学期管理中激活一个学期。", "warning")
		return None
	return semester


@students_bp.route("/")
@login_required
def list_students():
	semester = _require_semester_redirect()
	if semester is None:
		return redirect(url_for("semester_bp.list_semesters"))

	stage = request.args.get("stage", "全部")
	group_number = request.args.get("group", "")
	keyword = (request.args.get("q") or "").strip()

	query = Student.query.filter_by(semester_id=semester.id)
	if stage != "全部":
		query = query.filter_by(stage=stage)
	if group_number:
		try:
			query = query.filter_by(group_number=int(group_number))
		except ValueError:
			pass
	if keyword:
		query = query.filter((Student.name.contains(keyword)) | (Student.student_id.contains(keyword)))

	students = query.order_by(Student.group_number.asc().nullsfirst(), Student.student_id.asc()).all()
	groups = sorted({s.group_number for s in Student.query.filter_by(semester_id=semester.id).all() if s.group_number})

	if request.args.get("export") == "1":
		wb = Workbook()
		ws = wb.active
		ws.append(["姓名", "学号", "支部", "阶段", "小组", "状态"])
		for s in students:
			ws.append([s.name, s.student_id, s.department, s.stage, s.group_number or "", s.status])
		stream = BytesIO()
		wb.save(stream)
		stream.seek(0)
		return send_file(stream, as_attachment=True, download_name="学员列表.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
	return render_template(
		"students/list.html",
		students=students,
		groups=groups,
		stage=stage,
		group_number=group_number,
		keyword=keyword,
		semester=semester,
	)


@students_bp.route("/new", methods=["GET", "POST"])
@login_required
def new_student():
	semester = _require_semester_redirect()
	if semester is None:
		return redirect(url_for("semester_bp.list_semesters"))

	if request.method == "POST":
		name = (request.form.get("name") or "").strip()
		sid = (request.form.get("student_id") or "").strip()
		stage = (request.form.get("stage") or "积极分子").strip()
		if not name or not sid:
			flash("姓名和学号为必填项。", "danger")
			return render_template("students/form.html", student=None)
		if stage not in ["积极分子", "发展对象", "预备党员"]:
			flash("阶段取值不正确。", "danger")
			return render_template("students/form.html", student=None)
		if Student.query.filter_by(semester_id=semester.id, student_id=sid, stage=stage).first():
			flash("该学号在当前学期已存在。", "danger")
			return render_template("students/form.html", student=None)

		student = Student(
			semester_id=semester.id,
			name=name,
			student_id=sid,
			department=(request.form.get("department") or "").strip(),
			contact=(request.form.get("contact") or "").strip(),
			stage=stage,
			status="在读",
		)
		db.session.add(student)
		db.session.flush()
		_ensure_student_assignments(student, semester)
		db.session.commit()
		flash("学员已添加。", "success")
		return redirect(url_for("students_bp.list_students"))
	return render_template("students/form.html", student=None)


@students_bp.route("/<int:student_id>/edit", methods=["GET", "POST"])
@login_required
def edit_student(student_id: int):
	student = db.session.get(Student, student_id)
	if not student:
		flash("学员不存在。", "danger")
		return redirect(url_for("students_bp.list_students"))

	if request.method == "POST":
		name = (request.form.get("name") or "").strip()
		sid = (request.form.get("student_id") or "").strip()
		if not name or not sid:
			flash("姓名和学号为必填项。", "danger")
			return render_template("students/form.html", student=student)
		stage = (request.form.get("stage") or student.stage).strip()
		dup = Student.query.filter_by(semester_id=student.semester_id, student_id=sid, stage=stage).first()
		if dup and dup.id != student.id:
			flash("该学号在当前学期同阶段已存在。", "danger")
			return render_template("students/form.html", student=student)
		student.name = name
		student.student_id = sid
		student.department = (request.form.get("department") or "").strip()
		student.contact = (request.form.get("contact") or "").strip()
		student.stage = stage
		student.status = (request.form.get("status") or student.status).strip()
		db.session.commit()
		flash("学员信息已更新。", "success")
		return redirect(url_for("students_bp.list_students"))
	return render_template("students/form.html", student=student)


@students_bp.route("/<int:student_id>/delete", methods=["POST"])
@login_required
def delete_student(student_id: int):
	student = db.session.get(Student, student_id)
	if not student:
		flash("学员不存在。", "danger")
		return redirect(url_for("students_bp.list_students"))
	db.session.delete(student)
	db.session.commit()
	flash("学员已删除。", "success")
	return redirect(url_for("students_bp.list_students"))


@students_bp.route("/import", methods=["GET", "POST"])
@login_required
def import_students():
	semester = _require_semester_redirect()
	if semester is None:
		return redirect(url_for("semester_bp.list_semesters"))

	result = None
	if request.method == "POST":
		file = request.files.get("file")
		if not file:
			flash("请上传 xlsx 文件。", "danger")
			return redirect(url_for("students_bp.import_students"))

		workbook = load_workbook(file, read_only=True)
		sheet = workbook.active
		success_count = 0
		failed = []
		for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
			name, sid, department, stage = (row + (None, None, None, None))[:4]
			if not name or not sid:
				failed.append(f"第{idx}行: 姓名或学号为空")
				continue
			stage_value = (str(stage).strip() if stage else "积极分子")
			exists = Student.query.filter_by(semester_id=semester.id, student_id=str(sid).strip(), stage=stage_value).first()
			if exists:
				failed.append(f"第{idx}行: 学号 {sid} 在同阶段已存在")
				continue
			student = Student(
				semester_id=semester.id,
				name=str(name).strip(),
				student_id=str(sid).strip(),
				department=(str(department).strip() if department else ""),
				stage=stage_value,
				status="在读",
			)
			db.session.add(student)
			db.session.flush()
			_ensure_student_assignments(student, semester)
			success_count += 1
		db.session.commit()
		result = {"success": success_count, "failed": failed}
	return render_template("students/import.html", result=result)


@students_bp.route("/template")
@login_required
def download_template():
	wb = Workbook()
	ws = wb.active
	ws.title = "学员导入模板"
	ws.append(["姓名", "学号", "支部", "阶段"])

	stream = BytesIO()
	wb.save(stream)
	stream.seek(0)
	return send_file(
		stream,
		as_attachment=True,
		download_name="学员导入模板.xlsx",
		mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
	)


@students_bp.route("/grouping", methods=["GET", "POST"])
@login_required
def grouping():
	semester = _require_semester_redirect()
	if semester is None:
		return redirect(url_for("semester_bp.list_semesters"))

	students = Student.query.filter_by(semester_id=semester.id).order_by(Student.student_id.asc()).all()

	if request.method == "POST":
		action = request.form.get("action")
		if action == "auto":
			group_count = int(request.form.get("group_count") or 1)
			group_count = max(group_count, 1)
			for idx, student in enumerate(students):
				student.group_number = idx % group_count + 1
				student.is_group_leader = False
			db.session.commit()
			flash("已按学号顺序自动分组。", "success")
		else:
			for student in students:
				g = request.form.get(f"group_{student.id}")
				student.group_number = int(g) if g and g.isdigit() else None
				student.is_group_leader = request.form.get(f"leader_{student.id}") == "on"
			db.session.commit()
			flash("分组信息已保存。", "success")
		return redirect(url_for("students_bp.grouping"))

	return render_template("students/grouping.html", students=students)


@students_bp.route("/<int:student_id>/detail")
@login_required
def student_detail(student_id: int):
	student = db.session.get(Student, student_id)
	if not student:
		flash("学员不存在。", "danger")
		return redirect(url_for("students_bp.list_students"))

	attendance_records = (
		Attendance.query.filter_by(student_id=student.id)
		.join(CourseSession, Attendance.session_id == CourseSession.id)
		.order_by(CourseSession.session_number.asc())
		.all()
	)
	assignment_records = Assignment.query.filter_by(student_id=student.id).order_by(Assignment.id.desc()).all()
	volunteer_records = VolunteerRecord.query.filter_by(student_id=student.id).order_by(VolunteerRecord.id.desc()).all()
	exam_records = ExamRecord.query.filter_by(student_id=student.id).order_by(ExamRecord.id.desc()).all()
	final_score = FinalScore.query.filter_by(student_id=student.id, semester_id=student.semester_id).first()

	return render_template(
		"students/detail.html",
		student=student,
		attendance_records=attendance_records,
		assignment_records=assignment_records,
		volunteer_records=volunteer_records,
		exam_records=exam_records,
		final_score=final_score,
		now=datetime.utcnow(),
	)