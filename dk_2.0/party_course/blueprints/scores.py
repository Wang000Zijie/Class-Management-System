from datetime import datetime
from io import BytesIO

from flask import Blueprint, flash, redirect, render_template, request, send_file, url_for
from flask_login import login_required
from openpyxl import Workbook, load_workbook

from models import Assignment, Attendance, CourseSession, ExamRecord, FinalScore, Semester, Student, VolunteerRecord, db


scores_bp = Blueprint("scores_bp", __name__)


def _current_semester():
	return Semester.query.filter_by(status="active").order_by(Semester.id.desc()).first()


def _stage_weights(stage: str):
	if stage == "预备党员":
		return {"exam": 30, "assignment": 20, "volunteer": 50}
	if stage == "积极分子":
		return {"exam": 40, "assignment": 20, "volunteer": 40}
	# 发展对象未单独给出时，默认沿用积极分子配置
	return {"exam": 40, "assignment": 20, "volunteer": 40}


def _stage_period(semester: Semester, stage: str) -> str:
	if stage == "积极分子":
		return (semester.period_positive or "").strip()
	if stage == "发展对象":
		return (semester.period_development or "").strip()
	if stage == "预备党员":
		return (semester.period_probationary or "").strip()
	return ""


def _attendance_metrics(semester_id: int, student_id: int, total_sessions: int):
	if total_sessions <= 0:
		return {
			"attended": 0,
			"absent": 0,
			"display": "0/0",
			"ok": True,
			"threshold": 0,
		}

	attend_count = (
		Attendance.query.join(CourseSession, Attendance.session_id == CourseSession.id)
		.filter(CourseSession.semester_id == semester_id, Attendance.student_id == student_id)
		.filter(Attendance.status.in_(["到场", "线上"]))
		.count()
	)
	leave_count = (
		Attendance.query.join(CourseSession, Attendance.session_id == CourseSession.id)
		.filter(CourseSession.semester_id == semester_id, Attendance.student_id == student_id, Attendance.status == "请假")
		.count()
	)
	absent_count = total_sessions - attend_count - leave_count
	absent_count = max(absent_count, 0)
	threshold = total_sessions // 3
	ok = absent_count <= threshold
	return {
		"attended": attend_count,
		"absent": absent_count,
		"display": f"{attend_count}/{total_sessions}",
		"ok": ok,
		"threshold": threshold,
	}


def _recalculate_final_scores(semester: Semester, commit: bool = True):
	students = Student.query.filter_by(semester_id=semester.id).all()
	total_sessions = CourseSession.query.filter_by(semester_id=semester.id).count()
	for stu in students:
		attendance = _attendance_metrics(semester.id, stu.id, total_sessions)
		attendance_score = attendance["attended"]
		attendance_ok = attendance["ok"]

		tasks = Assignment.query.filter_by(student_id=stu.id, semester_id=semester.id).count()
		passed_tasks = Assignment.query.filter_by(student_id=stu.id, semester_id=semester.id, status="已通过").count()
		assignment_score = (passed_tasks / tasks * 100) if tasks else 0

		hours = db.session.query(db.func.sum(VolunteerRecord.hours)).filter_by(student_id=stu.id, semester_id=semester.id).scalar() or 0
		if semester.volunteer_target_hours > 0:
			volunteer_score = min(hours / semester.volunteer_target_hours, 1) * 100
		else:
			volunteer_score = 100 if hours > 0 else 0

		exam = ExamRecord.query.filter_by(student_id=stu.id, semester_id=semester.id).first()
		exam_score = exam.score if exam else 0

		weights = _stage_weights(stu.stage)
		total_score = (
			exam_score * weights["exam"]
			+ assignment_score * weights["assignment"]
			+ volunteer_score * weights["volunteer"]
		) / 100

		final = FinalScore.query.filter_by(student_id=stu.id, semester_id=semester.id).first()
		if not final:
			final = FinalScore(student_id=stu.id, semester_id=semester.id)
			db.session.add(final)
		final.attendance_score = round(attendance_score, 2)
		final.assignment_score = round(assignment_score, 2)
		final.volunteer_score = round(volunteer_score, 2)
		final.exam_score = round(exam_score, 2)
		final.total_score = round(total_score, 2)
		final.is_passed = attendance_ok and (stu.status != "取消资格") and (total_score >= semester.pass_threshold)

		if final.cert_number:
			if final.is_passed:
				if final.cert_number.endswith("(错误)"):
					final.cert_number = final.cert_number[:-4]
				elif final.cert_number.endswith("（错误）"):
					final.cert_number = final.cert_number[:-4]
			else:
				if not (final.cert_number.endswith("(错误)") or final.cert_number.endswith("（错误）")):
					final.cert_number = f"{final.cert_number}(错误)"

	if commit:
		db.session.commit()


@scores_bp.route("/exam", methods=["GET", "POST"])
@login_required
def exam_input():
	semester = _current_semester()
	if not semester:
		flash("请先激活学期。", "warning")
		return redirect(url_for("semester_bp.list_semesters"))
	students = Student.query.filter_by(semester_id=semester.id).order_by(Student.student_id.asc()).all()

	if request.method == "POST":
		if request.files.get("file"):
			wb = load_workbook(request.files["file"], read_only=True)
			ws = wb.active
			for row in ws.iter_rows(min_row=2, values_only=True):
				sid, score = (row + (None, None))[:2]
				stu = Student.query.filter_by(semester_id=semester.id, student_id=str(sid).strip()).first() if sid else None
				if not stu:
					continue
				rec = ExamRecord.query.filter_by(student_id=stu.id, semester_id=semester.id).first()
				if not rec:
					rec = ExamRecord(student_id=stu.id, semester_id=semester.id)
					db.session.add(rec)
				val = float(score or 0)
				if val < 0 or val > 100:
					continue
				rec.score = val
				rec.exam_time = datetime.utcnow()
		else:
			for stu in students:
				score = request.form.get(f"score_{stu.id}")
				cheating = request.form.get(f"cheating_{stu.id}") == "on"
				note = (request.form.get(f"cheat_note_{stu.id}") or "").strip()
				if score is None or score == "":
					continue
				rec = ExamRecord.query.filter_by(student_id=stu.id, semester_id=semester.id).first()
				if not rec:
					rec = ExamRecord(student_id=stu.id, semester_id=semester.id)
					db.session.add(rec)
				val = float(score)
				if val < 0 or val > 100:
					flash(f"{stu.name} 的分数超出范围，已跳过。", "danger")
					continue
				rec.score = val
				rec.exam_time = datetime.utcnow()
				rec.is_cheating = cheating
				rec.cheating_note = note
				if cheating:
					rec.score = 0
					stu.status = "取消资格"
					stu.disqualified_reason = note or "考试作弊"
		db.session.commit()
		flash("考试成绩已保存。", "success")
		return redirect(url_for("scores_bp.exam_input"))

	exam_map = {r.student_id: r for r in ExamRecord.query.filter_by(semester_id=semester.id).all()}
	return render_template("scores/exam.html", students=students, exam_map=exam_map)


@scores_bp.route("/calculate", methods=["POST"])
@login_required
def calculate_scores():
	semester = _current_semester()
	if not semester:
		flash("请先激活学期。", "warning")
		return redirect(url_for("semester_bp.list_semesters"))

	_recalculate_final_scores(semester)
	flash("综合成绩已计算完成。", "success")
	return redirect(url_for("scores_bp.result_overview"))


@scores_bp.route("/result")
@login_required
def result_overview():
	active_semester = _current_semester()
	if not active_semester:
		flash("请先激活学期。", "warning")
		return redirect(url_for("semester_bp.list_semesters"))

	semester_id = request.args.get("semester_id", type=int) or active_semester.id
	semester = db.session.get(Semester, semester_id)
	if not semester:
		flash("学期不存在。", "danger")
		return redirect(url_for("scores_bp.result_overview"))

	# Keep overview consistent with latest student/assignment/exam/volunteer changes.
	_recalculate_final_scores(semester)

	stage = request.args.get("stage", "全部")
	period = (request.args.get("period") or "全部").strip()
	sort_order = (request.args.get("sort") or "desc").strip().lower()
	if sort_order not in ["desc", "asc"]:
		sort_order = "desc"

	query = FinalScore.query.join(Student, FinalScore.student_id == Student.id).filter(FinalScore.semester_id == semester.id)
	if stage != "全部":
		query = query.filter(Student.stage == stage)
	if sort_order == "asc":
		query = query.order_by(FinalScore.total_score.asc(), Student.name.asc())
	else:
		query = query.order_by(FinalScore.total_score.desc(), Student.name.asc())

	total_sessions = CourseSession.query.filter_by(semester_id=semester.id).count()
	raw_rows = query.all()
	rows = []
	for item in raw_rows:
		stage_period = _stage_period(semester, item.student.stage)
		if period != "全部" and stage_period != period:
			continue
		attendance = _attendance_metrics(semester.id, item.student_id, total_sessions)
		rows.append(
			{
				"final": item,
				"stage_period": stage_period,
				"attendance_display": attendance["display"],
				"attendance_ok": attendance["ok"],
				"attendance_absent": attendance["absent"],
			}
		)

	semesters = Semester.query.order_by(Semester.year.desc(), Semester.id.desc()).all()
	period_options = []
	for p in [semester.period_positive, semester.period_development, semester.period_probationary]:
		v = (p or "").strip()
		if v and v not in period_options:
			period_options.append(v)
	return render_template(
		"scores/result.html",
		rows=rows,
		stage=stage,
		period=period,
		threshold=semester.pass_threshold,
		semester_id=semester.id,
		semesters=semesters,
		period_options=period_options,
		sort_order=sort_order,
	)


@scores_bp.route("/issue_certs", methods=["POST"])
@login_required
def issue_certs():
	semester = _current_semester()
	if not semester:
		flash("请先激活学期。", "warning")
		return redirect(url_for("scores_bp.result_overview"))

	passed = FinalScore.query.filter_by(semester_id=semester.id, is_passed=True).order_by(FinalScore.total_score.desc()).all()
	counters = {"积极分子": 0, "预备党员": 0, "发展对象": 0}
	for item in passed:
		stu = db.session.get(Student, item.student_id)
		stage = stu.stage if stu else "积极分子"
		period = _stage_period(semester, stage) or str(semester.year)
		if stage not in counters:
			counters[stage] = 0
		counters[stage] += 1
		seq = f"{counters[stage]:03d}"
		if stage == "积极分子":
			item.cert_number = f"T{period}{seq}"
		elif stage == "预备党员":
			item.cert_number = f"{period}T{seq}"
		elif stage == "发展对象":
			item.cert_number = f"{period}{seq}T"
		else:
			item.cert_number = f"{period}{seq}"
		item.cert_issued_at = datetime.utcnow()
	db.session.commit()
	flash("证书编号已发放。", "success")
	return redirect(url_for("scores_bp.result_overview"))


@scores_bp.route("/publish")
def publish_page():
	semester = _current_semester()
	if not semester:
		return "暂无公示数据"
	rows = (
		FinalScore.query.join(Student, FinalScore.student_id == Student.id)
		.filter(FinalScore.semester_id == semester.id, FinalScore.is_passed == True)
		.order_by(Student.stage.asc(), Student.name.asc())
		.all()
	)
	return render_template("scores/publish.html", rows=rows, semester=semester)


@scores_bp.route("/export")
@login_required
def export_all():
	semester = _current_semester()
	if not semester:
		flash("请先激活学期。", "warning")
		return redirect(url_for("scores_bp.result_overview"))

	wb = Workbook()
	ws = wb.active
	ws.append(["姓名", "学号", "党支部", "阶段", "出勤次数", "考试分数", "综合得分", "是否通过", "证书编号", "志愿时长"])

	students = Student.query.filter_by(semester_id=semester.id).order_by(Student.student_id.asc()).all()
	total_sessions = CourseSession.query.filter_by(semester_id=semester.id).count()
	for stu in students:
		attendance = (
			Attendance.query.join(CourseSession, Attendance.session_id == CourseSession.id)
			.filter(CourseSession.semester_id == semester.id, Attendance.student_id == stu.id)
			.filter(Attendance.status.in_(["到场", "线上"]))
			.count()
		)
		exam = ExamRecord.query.filter_by(student_id=stu.id, semester_id=semester.id).first()
		fs = FinalScore.query.filter_by(student_id=stu.id, semester_id=semester.id).first()
		vol_hours = db.session.query(db.func.sum(VolunteerRecord.hours)).filter_by(student_id=stu.id, semester_id=semester.id).scalar() or 0
		ws.append([
			stu.name,
			stu.student_id,
			stu.department,
			stu.stage,
			f"{attendance}/{total_sessions}",
			exam.score if exam else "",
			fs.total_score if fs else "",
			"通过" if fs and fs.is_passed else "未通过",
			fs.cert_number if fs else "",
			round(vol_hours, 2),
		])

	stream = BytesIO()
	wb.save(stream)
	stream.seek(0)
	return send_file(stream, as_attachment=True, download_name="成绩归档.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")