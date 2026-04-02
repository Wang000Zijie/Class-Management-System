from datetime import datetime
from io import BytesIO

from flask import Blueprint, flash, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from openpyxl import Workbook, load_workbook

from models import Semester, Student, VolunteerRecord, db


volunteers_bp = Blueprint("volunteers_bp", __name__)


def _current_semester():
	return Semester.query.filter_by(status="active").order_by(Semester.id.desc()).first()


@volunteers_bp.route("/")
@login_required
def volunteer_overview():
	semester = _current_semester()
	if not semester:
		flash("请先激活学期。", "warning")
		return redirect(url_for("semester_bp.list_semesters"))

	stage = request.args.get("stage", "全部")
	period = (request.args.get("period") or "全部").strip()
	query = Student.query.filter_by(semester_id=semester.id)
	if stage != "全部":
		query = query.filter_by(stage=stage)
	students = query.order_by(Student.student_id.asc()).all()

	rows = []
	for stu in students:
		stu_period = semester.period_for_stage(stu.stage)
		if period != "全部" and stu_period != period:
			continue
		records = VolunteerRecord.query.filter_by(student_id=stu.id, semester_id=semester.id).all()
		total_hours = round(sum(r.hours for r in records), 2)
		verified = all(r.verified for r in records) if records else False
		rows.append({"student": stu, "hours": total_hours, "verified": verified, "qualified": semester.volunteer_target_hours <= 0 or total_hours >= semester.volunteer_target_hours, "stage_period": stu_period})
	if request.args.get("export") == "1":
		wb = Workbook()
		ws = wb.active
		ws.append(["姓名", "学号", "阶段", "累计志愿时长", "核实状态", "是否达标"])
		for row in rows:
			ws.append([
				row["student"].name,
				row["student"].student_id,
				row["student"].stage,
				row["hours"],
				"已核实" if row["verified"] else "未核实",
				"达标" if row["qualified"] else "未达标",
			])
		stream = BytesIO()
		wb.save(stream)
		stream.seek(0)
		return send_file(stream, as_attachment=True, download_name="志愿时长总览.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

	period_options = []
	for p in [semester.period_positive, semester.period_development, semester.period_probationary]:
		v = (p or "").strip()
		if v and v not in period_options:
			period_options.append(v)
	return render_template("volunteers/overview.html", rows=rows, stage=stage, period=period, period_options=period_options, semester=semester)


@volunteers_bp.route("/student/<int:student_id>", methods=["GET", "POST"])
@login_required
def student_volunteer(student_id: int):
	student = db.session.get(Student, student_id)
	if not student:
		flash("学员不存在。", "danger")
		return redirect(url_for("volunteers_bp.volunteer_overview"))

	semester_id = student.semester_id
	if request.method == "POST":
		action = request.form.get("action")
		if action == "delete":
			rec = db.session.get(VolunteerRecord, request.form.get("record_id", type=int))
			if rec and rec.student_id == student.id:
				db.session.delete(rec)
		else:
			rec_id = request.form.get("record_id", type=int)
			rec = db.session.get(VolunteerRecord, rec_id) if rec_id else VolunteerRecord(student_id=student.id, semester_id=semester_id)
			if not rec_id:
				db.session.add(rec)
			activity_name = (request.form.get("activity_name") or "").strip()
			hours = request.form.get("hours", type=float)
			if not activity_name:
				flash("活动名称为必填项。", "danger")
				return redirect(url_for("volunteers_bp.student_volunteer", student_id=student.id))
			if hours is None or hours < 0:
				flash("时长必须为大于等于 0 的数字。", "danger")
				return redirect(url_for("volunteers_bp.student_volunteer", student_id=student.id))
			rec.activity_name = activity_name
			rec.hours = hours
			rec.proof_note = (request.form.get("proof_note") or "").strip()
			rec.verified = request.form.get("verified") == "on"
			rec.verified_by = current_user.username if rec.verified else ""
			rec.verified_at = datetime.utcnow() if rec.verified else None
		db.session.commit()
		flash("志愿记录保存成功。", "success")
		return redirect(url_for("volunteers_bp.student_volunteer", student_id=student.id))

	records = VolunteerRecord.query.filter_by(student_id=student.id, semester_id=semester_id).order_by(VolunteerRecord.id.desc()).all()
	total_hours = round(sum(r.hours for r in records), 2)
	return render_template("volunteers/student.html", student=student, records=records, total_hours=total_hours)


@volunteers_bp.route("/batch", methods=["GET", "POST"])
@login_required
def batch_volunteer():
	semester = _current_semester()
	if not semester:
		flash("请先激活学期。", "warning")
		return redirect(url_for("volunteers_bp.volunteer_overview"))

	if request.method == "POST":
		file = request.files.get("file")
		if not file:
			flash("请上传文件。", "danger")
			return redirect(url_for("volunteers_bp.batch_volunteer"))
		wb = load_workbook(file, read_only=True)
		ws = wb.active
		success, fail = 0, 0
		for row in ws.iter_rows(min_row=2, values_only=True):
			sid, activity, hours, proof = (row + (None, None, None, None))[:4]
			stu = Student.query.filter_by(semester_id=semester.id, student_id=str(sid).strip()).first() if sid else None
			if not stu or not activity:
				fail += 1
				continue
			db.session.add(
				VolunteerRecord(
					student_id=stu.id,
					semester_id=semester.id,
					activity_name=str(activity).strip(),
					hours=float(hours or 0),
					proof_note=str(proof or "").strip(),
					verified=False,
				)
			)
			success += 1
		db.session.commit()
		flash(f"导入完成：成功{success}，失败{fail}。", "info")
		return redirect(url_for("volunteers_bp.batch_volunteer"))
	return render_template("volunteers/batch.html")


@volunteers_bp.route("/summary")
@login_required
def volunteer_summary():
	semester = _current_semester()
	if not semester:
		flash("请先激活学期。", "warning")
		return redirect(url_for("volunteers_bp.volunteer_overview"))

	wb = Workbook()
	ws = wb.active
	ws.append(["姓名", "学号", "阶段", "累计志愿时长", "核实状态"])
	students = Student.query.filter_by(semester_id=semester.id).order_by(Student.student_id.asc()).all()
	for stu in students:
		records = VolunteerRecord.query.filter_by(student_id=stu.id, semester_id=semester.id).all()
		total_hours = round(sum(r.hours for r in records), 2)
		status = "已核实" if records and all(r.verified for r in records) else "未核实"
		ws.append([stu.name, stu.student_id, stu.stage, total_hours, status])
	stream = BytesIO()
	wb.save(stream)
	stream.seek(0)
	return send_file(stream, as_attachment=True, download_name="志愿时长汇总.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")