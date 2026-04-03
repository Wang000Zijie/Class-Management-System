from datetime import datetime
from io import BytesIO

from flask import Blueprint, flash, jsonify, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from openpyxl import Workbook, load_workbook

from models import Attendance, CourseSession, Semester, Student, db


attendance_bp = Blueprint("attendance_bp", __name__)


def _current_semester():
	return Semester.query.filter_by(status="active").order_by(Semester.id.desc()).first()


@attendance_bp.route("/")
@login_required
def attendance_overview():
	semester = _current_semester()
	if not semester:
		flash("请先激活学期。", "warning")
		return redirect(url_for("semester_bp.list_semesters"))

	selected_id = request.args.get("session_id", type=int)
	stage = request.args.get("stage", "全部")
	sessions = CourseSession.query.filter_by(semester_id=semester.id).order_by(CourseSession.session_number.asc()).all()

	rows = []
	if selected_id:
		students_query = Student.query.filter_by(semester_id=semester.id)
		if stage != "全部":
			students_query = students_query.filter_by(stage=stage)
		students = students_query.order_by(Student.student_id.asc()).all()
		att_map = {item.student_id: item for item in Attendance.query.filter_by(session_id=selected_id).all()}
		for stu in students:
			rows.append({"student": stu, "status": att_map.get(stu.id).status if att_map.get(stu.id) else "未记录"})

	if request.args.get("export") == "1" and selected_id:
		wb = Workbook()
		ws = wb.active
		ws.append(["姓名", "学号", "状态"])
		for row in rows:
			ws.append([row["student"].name, row["student"].student_id, row["status"]])
		stream = BytesIO()
		wb.save(stream)
		stream.seek(0)
		return send_file(stream, as_attachment=True, download_name="签到总览.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

	return render_template("attendance/overview.html", sessions=sessions, selected_id=selected_id, rows=rows, stage=stage)


@attendance_bp.route("/session/<int:session_id>", methods=["GET", "POST"])
@login_required
def session_checkin(session_id: int):
	session = db.session.get(CourseSession, session_id)
	if not session:
		flash("课次不存在。", "danger")
		return redirect(url_for("attendance_bp.attendance_overview"))

	if request.method == "POST":
		student_id = request.form.get("student_id", type=int)
		status = request.form.get("status")
		leave_reason = (request.form.get("leave_reason") or "").strip()
		if status not in ["到场", "线上", "请假", "缺席"]:
			return jsonify({"ok": False, "msg": "状态无效"}), 400
		if status == "请假" and not leave_reason:
			return jsonify({"ok": False, "msg": "请假需填写原因"}), 400

		rec = Attendance.query.filter_by(student_id=student_id, session_id=session_id).first()
		if not rec:
			rec = Attendance(student_id=student_id, session_id=session_id)
			db.session.add(rec)
		rec.status = status
		rec.leave_reason = leave_reason if status == "请假" else ""
		rec.checked_by = current_user.username
		rec.checked_at = datetime.utcnow()
		db.session.commit()
		return jsonify({"ok": True})

	stage = request.args.get("stage", "全部")
	students_query = Student.query.filter_by(semester_id=session.semester_id)
	if stage != "全部":
		students_query = students_query.filter_by(stage=stage)
	students = students_query.order_by(Student.group_number.asc().nullsfirst(), Student.student_id.asc()).all()

	records = {item.student_id: item for item in Attendance.query.filter_by(session_id=session_id).all()}
	rows = []
	count_map = {"到场": 0, "线上": 0, "请假": 0, "缺席": 0}
	for stu in students:
		rec = records.get(stu.id)
		status = rec.status if rec else "缺席"
		if status in count_map:
			count_map[status] += 1
		rows.append({"student": stu, "status": status, "leave_reason": rec.leave_reason if rec else ""})

	return render_template("attendance/session.html", session=session, rows=rows, stage=stage, count_map=count_map)


@attendance_bp.route("/session/<int:session_id>/mark_all_present", methods=["POST"])
@login_required
def mark_all_present(session_id: int):
	session = db.session.get(CourseSession, session_id)
	if not session:
		flash("课次不存在。", "danger")
		return redirect(url_for("attendance_bp.attendance_overview"))

	students = Student.query.filter_by(semester_id=session.semester_id).all()
	for stu in students:
		rec = Attendance.query.filter_by(student_id=stu.id, session_id=session_id).first()
		if not rec:
			rec = Attendance(student_id=stu.id, session_id=session_id)
			db.session.add(rec)
		rec.status = "到场"
		rec.leave_reason = ""
		rec.checked_by = current_user.username
		rec.checked_at = datetime.utcnow()
	db.session.commit()
	flash("已批量标记为到场。", "success")
	return redirect(url_for("attendance_bp.session_checkin", session_id=session_id))


@attendance_bp.route("/summary")
@login_required
def attendance_summary():
	semester = _current_semester()
	if not semester:
		flash("请先激活学期。", "warning")
		return redirect(url_for("semester_bp.list_semesters"))

	sessions = CourseSession.query.filter_by(semester_id=semester.id).order_by(CourseSession.session_number.asc()).all()
	students = Student.query.filter_by(semester_id=semester.id).order_by(Student.student_id.asc()).all()
	recs = Attendance.query.join(Student, Attendance.student_id == Student.id).filter(Student.semester_id == semester.id).all()

	map_status = {(r.student_id, r.session_id): r.status for r in recs}
	table_rows = []
	for stu in students:
		statuses = []
		valid_count = 0
		for sess in sessions:
			s = map_status.get((stu.id, sess.id), "未记录")
			statuses.append(s)
			if s in ["到场", "线上"]:
				valid_count += 1
		table_rows.append({"student": stu, "statuses": statuses, "valid_count": valid_count, "warn": valid_count < semester.min_attendance_sessions})

	if request.args.get("export") == "1":
		wb = Workbook()
		ws = wb.active
		ws.append(["姓名", "学号", *[f"第{s.session_number}讲" for s in sessions], "出勤次数"])
		for row in table_rows:
			ws.append([row["student"].name, row["student"].student_id, *row["statuses"], row["valid_count"]])
		stream = BytesIO()
		wb.save(stream)
		stream.seek(0)
		return send_file(stream, as_attachment=True, download_name="出勤汇总.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

	return render_template("attendance/summary.html", sessions=sessions, table_rows=table_rows, semester=semester)


@attendance_bp.route("/disqualify/<int:student_id>", methods=["POST"])
@login_required
def disqualify_student(student_id: int):
	reason = (request.form.get("reason") or "").strip()
	student = db.session.get(Student, student_id)
	if not student:
		flash("学员不存在。", "danger")
		return redirect(url_for("attendance_bp.attendance_summary"))
	student.status = "取消资格"
	student.disqualified_reason = reason or "出勤不足"
	db.session.commit()
	flash("已取消该学员资格。", "warning")
	return redirect(url_for("attendance_bp.attendance_summary"))


@attendance_bp.route("/batch", methods=["POST"])
@login_required
def batch_import_attendance():
	semester = _current_semester()
	if not semester:
		flash("请先激活学期。", "warning")
		return redirect(url_for("attendance_bp.attendance_overview"))

	file = request.files.get("file")
	if not file:
		flash("请上传 Excel 文件。", "danger")
		return redirect(url_for("attendance_bp.attendance_overview"))

	wb = load_workbook(file, read_only=True)
	ws = wb.active
	success = 0
	fail = 0
	for row in ws.iter_rows(min_row=2, values_only=True):
		sid, session_no, status = row[:3]
		student = Student.query.filter_by(semester_id=semester.id, student_id=str(sid).strip()).first()
		session = CourseSession.query.filter_by(semester_id=semester.id, session_number=int(session_no)).first() if session_no else None
		if not student or not session or status not in ["到场", "线上", "请假", "缺席"]:
			fail += 1
			continue
		rec = Attendance.query.filter_by(student_id=student.id, session_id=session.id).first()
		if not rec:
			rec = Attendance(student_id=student.id, session_id=session.id)
			db.session.add(rec)
		rec.status = status
		rec.checked_by = current_user.username
		rec.checked_at = datetime.utcnow()
		success += 1
	db.session.commit()
	flash(f"批量导入完成：成功 {success}，失败 {fail}。", "info")
	return redirect(url_for("attendance_bp.attendance_overview"))