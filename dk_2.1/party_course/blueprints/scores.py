from collections import defaultdict
from datetime import datetime
from io import BytesIO

from flask import Blueprint, flash, redirect, render_template, request, send_file, url_for
from flask_login import login_required
from openpyxl import Workbook, load_workbook
from sqlalchemy import case

from models import Assignment, Attendance, CourseSession, ExamRecord, FinalScore, Semester, Student, VolunteerRecord, db


scores_bp = Blueprint("scores_bp", __name__)


def _current_semester():
	return Semester.query.filter_by(status="active").order_by(Semester.id.desc()).first()


def _stage_weights(stage: str):
	if stage == "预备党员":
		return {"exam": 30, "assignment": 20, "volunteer": 50}
	if stage == "积极分子":
		return {"exam": 40, "assignment": 20, "volunteer": 40}
	# 发展对象未单独给出时，默认沿用积极分子配置
	return {"exam": 40, "assignment": 20, "volunteer": 40}


def _stage_period(semester: Semester, stage: str) -> str:
	if stage == "积极分子":
		return (semester.period_positive or "").strip()
	if stage == "发展对象":
		return (semester.period_development or "").strip()
	if stage == "预备党员":
		return (semester.period_probationary or "").strip()
	return ""


def _attendance_count_maps(semester_id: int, student_ids: list[int]) -> tuple[dict[int, int], dict[int, int]]:
	present_map: dict[int, int] = defaultdict(int)
	leave_map: dict[int, int] = defaultdict(int)
	if not student_ids:
		return present_map, leave_map

	rows = (
		db.session.query(Attendance.student_id, Attendance.status, db.func.count(Attendance.id))
		.join(CourseSession, Attendance.session_id == CourseSession.id)
		.filter(CourseSession.semester_id == semester_id, Attendance.student_id.in_(student_ids))
		.group_by(Attendance.student_id, Attendance.status)
		.all()
	)
	for student_id, status, count_value in rows:
		count = int(count_value or 0)
		if status in {"到场", "线上"}:
			present_map[int(student_id)] += count
		elif status == "请假":
			leave_map[int(student_id)] += count

	return present_map, leave_map


def _attendance_metrics_map(semester_id: int, student_ids: list[int], total_sessions: int) -> dict[int, dict]:
	result: dict[int, dict] = {}
	if not student_ids:
		return result

	if total_sessions <= 0:
		for student_id in student_ids:
			result[student_id] = {
				"attended": 0,
				"absent": 0,
				"display": "0/0",
				"ok": True,
				"threshold": 0,
			}
		return result

	present_map, leave_map = _attendance_count_maps(semester_id, student_ids)
	threshold = total_sessions // 3
	for student_id in student_ids:
		attended = int(present_map.get(student_id, 0) or 0)
		leave = int(leave_map.get(student_id, 0) or 0)
		absent = max(total_sessions - attended - leave, 0)
		result[student_id] = {
			"attended": attended,
			"absent": absent,
			"display": f"{attended}/{total_sessions}",
			"ok": absent <= threshold,
			"threshold": threshold,
		}

	return result


def _attendance_metrics(semester_id: int, student_id: int, total_sessions: int):
	metrics = _attendance_metrics_map(semester_id, [student_id], total_sessions)
	if student_id in metrics:
		return metrics[student_id]
	return {
		"attended": 0,
		"absent": max(total_sessions, 0),
		"display": f"0/{max(total_sessions, 0)}",
		"ok": total_sessions <= 0,
		"threshold": total_sessions // 3 if total_sessions > 0 else 0,
	}


def _recalculate_final_scores(semester: Semester, commit: bool = True):
	students = Student.query.filter_by(semester_id=semester.id).all()
	student_ids = [stu.id for stu in students]
	total_sessions = CourseSession.query.filter_by(semester_id=semester.id).count()
	attendance_map = _attendance_metrics_map(semester.id, student_ids, total_sessions)

	assignment_stats: dict[int, dict[str, int]] = defaultdict(lambda: {"tasks": 0, "passed": 0})
	if student_ids:
		assignment_rows = (
			db.session.query(
				Assignment.student_id,
				db.func.count(Assignment.id),
				db.func.sum(case((Assignment.status == "已通过", 1), else_=0)),
			)
			.filter(Assignment.semester_id == semester.id, Assignment.student_id.in_(student_ids))
			.group_by(Assignment.student_id)
			.all()
		)
		for student_id, task_count, passed_count in assignment_rows:
			assignment_stats[int(student_id)] = {
				"tasks": int(task_count or 0),
				"passed": int(passed_count or 0),
			}

	volunteer_hours_map: dict[int, float] = {}
	if student_ids:
		volunteer_rows = (
			db.session.query(VolunteerRecord.student_id, db.func.sum(VolunteerRecord.hours))
			.filter(VolunteerRecord.semester_id == semester.id, VolunteerRecord.student_id.in_(student_ids))
			.group_by(VolunteerRecord.student_id)
			.all()
		)
		for student_id, total_hours in volunteer_rows:
			volunteer_hours_map[int(student_id)] = float(total_hours or 0)

	exam_map: dict[int, ExamRecord] = {}
	if student_ids:
		for exam in ExamRecord.query.filter(ExamRecord.semester_id == semester.id, ExamRecord.student_id.in_(student_ids)).all():
			exam_map[exam.student_id] = exam

	final_map: dict[int, FinalScore] = {}
	if student_ids:
		for item in FinalScore.query.filter(FinalScore.semester_id == semester.id, FinalScore.student_id.in_(student_ids)).all():
			final_map[item.student_id] = item

	for stu in students:
		attendance = attendance_map.get(stu.id)
		if attendance is None:
			attendance = _attendance_metrics(semester.id, stu.id, total_sessions)
		attendance_score = attendance["attended"]
		attendance_ok = attendance["ok"]

		student_assignment_stat = assignment_stats.get(stu.id, {"tasks": 0, "passed": 0})
		tasks = int(student_assignment_stat["tasks"])
		passed_tasks = int(student_assignment_stat["passed"])
		assignment_score = (passed_tasks / tasks * 100) if tasks else 0

		hours = float(volunteer_hours_map.get(stu.id, 0) or 0)
		if semester.volunteer_target_hours > 0:
			volunteer_score = min(hours / semester.volunteer_target_hours, 1) * 100
		else:
			volunteer_score = 100 if hours > 0 else 0

		exam = exam_map.get(stu.id)
		exam_score = exam.score if exam else 0

		weights = _stage_weights(stu.stage)
		total_score = (
			exam_score * weights["exam"]
			+ assignment_score * weights["assignment"]
			+ volunteer_score * weights["volunteer"]
		) / 100

		final = final_map.get(stu.id)
		if not final:
			final = FinalScore(student_id=stu.id, semester_id=semester.id)
			db.session.add(final)
			final_map[stu.id] = final
		final.attendance_score = round(attendance_score, 2)
		final.assignment_score = round(assignment_score, 2)
		final.volunteer_score = round(volunteer_score, 2)
		final.exam_score = round(exam_score, 2)
		final.total_score = round(total_score, 2)
		final.is_passed = attendance_ok and (stu.status != "取消资格") and (total_score >= semester.pass_threshold)

		if final.cert_number:
			if final.is_passed:
				if final.cert_number.endswith("(错误)"):
					final.cert_number = final.cert_number[:-4]
				elif final.cert_number.endswith("（错误）"):
					final.cert_number = final.cert_number[:-4]
			else:
				if not (final.cert_number.endswith("(错误)") or final.cert_number.endswith("（错误）")):
					final.cert_number = f"{final.cert_number}(错误)"

	if commit:
		db.session.commit()


@scores_bp.route("/exam", methods=["GET", "POST"])
@login_required
def exam_input():
	semester = _current_semester()
	if not semester:
		flash("请先激活学期。", "warning")
		return redirect(url_for("semester_bp.list_semesters"))
	students = Student.query.filter_by(semester_id=semester.id).order_by(Student.student_id.asc()).all()

	if request.method == "POST":
		if request.files.get("file"):
			wb = load_workbook(request.files["file"], read_only=True)
			ws = wb.active
			for row in ws.iter_rows(min_row=2, values_only=True):
				sid, score = (row + (None, None))[:2]
				stu = Student.query.filter_by(semester_id=semester.id, student_id=str(sid).strip()).first() if sid else None
				if not stu:
					continue
				rec = ExamRecord.query.filter_by(student_id=stu.id, semester_id=semester.id).first()
				if not rec:
					rec = ExamRecord(student_id=stu.id, semester_id=semester.id)
					db.session.add(rec)
				val = float(score or 0)
				if val < 0 or val > 100:
					continue
				rec.score = val
				rec.exam_time = datetime.utcnow()
		else:
			for stu in students:
				score = request.form.get(f"score_{stu.id}")
				cheating = request.form.get(f"cheating_{stu.id}") == "on"
				note = (request.form.get(f"cheat_note_{stu.id}") or "").strip()
				if score is None or score == "":
					continue
				rec = ExamRecord.query.filter_by(student_id=stu.id, semester_id=semester.id).first()
				if not rec:
					rec = ExamRecord(student_id=stu.id, semester_id=semester.id)
					db.session.add(rec)
				val = float(score)
				if val < 0 or val > 100:
					flash(f"{stu.name} 的分数超出范围，已跳过。", "danger")
					continue
				rec.score = val
				rec.exam_time = datetime.utcnow()
				rec.is_cheating = cheating
				rec.cheating_note = note
				if cheating:
					rec.score = 0
					stu.status = "取消资格"
					stu.disqualified_reason = note or "考试作弊"
		db.session.commit()
		flash("考试成绩已保存。", "success")
		return redirect(url_for("scores_bp.exam_input"))

	exam_map = {r.student_id: r for r in ExamRecord.query.filter_by(semester_id=semester.id).all()}
	return render_template("scores/exam.html", students=students, exam_map=exam_map)


@scores_bp.route("/calculate", methods=["POST"])
@login_required
def calculate_scores():
	semester = _current_semester()
	if not semester:
		flash("请先激活学期。", "warning")
		return redirect(url_for("semester_bp.list_semesters"))

	_recalculate_final_scores(semester)
	flash("综合成绩已计算完成。", "success")
	return redirect(url_for("scores_bp.result_overview"))


@scores_bp.route("/result")
@login_required
def result_overview():
	active_semester = _current_semester()
	if not active_semester:
		flash("请先激活学期。", "warning")
		return redirect(url_for("semester_bp.list_semesters"))

	semester_id = request.args.get("semester_id", type=int) or active_semester.id
	semester = db.session.get(Semester, semester_id)
	if not semester:
		flash("学期不存在。", "danger")
		return redirect(url_for("scores_bp.result_overview"))

	# Keep overview consistent with latest student/assignment/exam/volunteer changes.
	_recalculate_final_scores(semester)

	stage = request.args.get("stage", "全部")
	period = (request.args.get("period") or "全部").strip()
	sort_order = (request.args.get("sort") or "desc").strip().lower()
	if sort_order not in ["desc", "asc"]:
		sort_order = "desc"

	query = FinalScore.query.join(Student, FinalScore.student_id == Student.id).filter(FinalScore.semester_id == semester.id)
	if stage != "全部":
		query = query.filter(Student.stage == stage)
	if sort_order == "asc":
		query = query.order_by(FinalScore.total_score.asc(), Student.name.asc())
	else:
		query = query.order_by(FinalScore.total_score.desc(), Student.name.asc())

	total_sessions = CourseSession.query.filter_by(semester_id=semester.id).count()
	raw_rows = query.all()
	attendance_map = _attendance_metrics_map(semester.id, [item.student_id for item in raw_rows], total_sessions)
	rows = []
	for item in raw_rows:
		stage_period = _stage_period(semester, item.student.stage)
		if period != "全部" and stage_period != period:
			continue
		attendance = attendance_map.get(item.student_id)
		if attendance is None:
			attendance = _attendance_metrics(semester.id, item.student_id, total_sessions)
		rows.append(
			{
				"final": item,
				"stage_period": stage_period,
				"attendance_display": attendance["display"],
				"attendance_ok": attendance["ok"],
				"attendance_absent": attendance["absent"],
			}
		)

	semesters = Semester.query.order_by(Semester.year.desc(), Semester.id.desc()).all()
	period_options = []
	for p in [semester.period_positive, semester.period_development, semester.period_probationary]:
		v = (p or "").strip()
		if v and v not in period_options:
			period_options.append(v)
	return render_template(
		"scores/result.html",
		rows=rows,
		stage=stage,
		period=period,
		threshold=semester.pass_threshold,
		semester_id=semester.id,
		semesters=semesters,
		period_options=period_options,
		sort_order=sort_order,
	)


@scores_bp.route("/issue_certs", methods=["POST"])
@login_required
def issue_certs():
	semester = _current_semester()
	if not semester:
		flash("请先激活学期。", "warning")
		return redirect(url_for("scores_bp.result_overview"))

	passed = FinalScore.query.filter_by(semester_id=semester.id, is_passed=True).order_by(FinalScore.total_score.desc()).all()
	student_map: dict[int, Student] = {}
	student_ids = [item.student_id for item in passed]
	if student_ids:
		for stu in Student.query.filter(Student.id.in_(student_ids)).all():
			student_map[stu.id] = stu
	counters = {"积极分子": 0, "预备党员": 0, "发展对象": 0}
	for item in passed:
		stu = student_map.get(item.student_id)
		stage = stu.stage if stu else "积极分子"
		period = _stage_period(semester, stage) or str(semester.year)
		if stage not in counters:
			counters[stage] = 0
		counters[stage] += 1
		seq = f"{counters[stage]:03d}"
		if stage == "积极分子":
			item.cert_number = f"T{period}{seq}"
		elif stage == "预备党员":
			item.cert_number = f"{period}T{seq}"
		elif stage == "发展对象":
			item.cert_number = f"{period}{seq}T"
		else:
			item.cert_number = f"{period}{seq}"
		item.cert_issued_at = datetime.utcnow()
	db.session.commit()
	flash("证书编号已发放。", "success")
	return redirect(url_for("scores_bp.result_overview"))


@scores_bp.route("/publish")
def publish_page():
	semester = _current_semester()
	if not semester:
		return "暂无公示数据"
	rows = (
		FinalScore.query.join(Student, FinalScore.student_id == Student.id)
		.filter(FinalScore.semester_id == semester.id, FinalScore.is_passed == True)
		.order_by(Student.stage.asc(), Student.name.asc())
		.all()
	)
	return render_template("scores/publish.html", rows=rows, semester=semester)


@scores_bp.route("/export")
@login_required
def export_all():
	semester = _current_semester()
	if not semester:
		flash("请先激活学期。", "warning")
		return redirect(url_for("scores_bp.result_overview"))

	wb = Workbook()
	ws = wb.active
	ws.append(["姓名", "学号", "党支部", "阶段", "出勤次数", "考试分数", "综合得分", "是否通过", "证书编号", "志愿时长"])

	students = Student.query.filter_by(semester_id=semester.id).order_by(Student.student_id.asc()).all()
	student_ids = [stu.id for stu in students]
	total_sessions = CourseSession.query.filter_by(semester_id=semester.id).count()
	attendance_present_map, _leave_map = _attendance_count_maps(semester.id, student_ids)
	exam_map: dict[int, ExamRecord] = {}
	final_map: dict[int, FinalScore] = {}
	volunteer_hours_map: dict[int, float] = {}

	if student_ids:
		for exam in ExamRecord.query.filter(ExamRecord.semester_id == semester.id, ExamRecord.student_id.in_(student_ids)).all():
			exam_map[exam.student_id] = exam
		for fs in FinalScore.query.filter(FinalScore.semester_id == semester.id, FinalScore.student_id.in_(student_ids)).all():
			final_map[fs.student_id] = fs
		for student_id, total_hours in (
			db.session.query(VolunteerRecord.student_id, db.func.sum(VolunteerRecord.hours))
			.filter(VolunteerRecord.semester_id == semester.id, VolunteerRecord.student_id.in_(student_ids))
			.group_by(VolunteerRecord.student_id)
			.all()
		):
			volunteer_hours_map[int(student_id)] = float(total_hours or 0)

	for stu in students:
		attendance = int(attendance_present_map.get(stu.id, 0) or 0)
		exam = exam_map.get(stu.id)
		fs = final_map.get(stu.id)
		vol_hours = float(volunteer_hours_map.get(stu.id, 0) or 0)
		ws.append([
			stu.name,
			stu.student_id,
			stu.department,
			stu.stage,
			f"{attendance}/{total_sessions}",
			exam.score if exam else "",
			fs.total_score if fs else "",
			"通过" if fs and fs.is_passed else "未通过",
			fs.cert_number if fs else "",
			round(vol_hours, 2),
		])

	stream = BytesIO()
	wb.save(stream)
	stream.seek(0)
	return send_file(stream, as_attachment=True, download_name="成绩归档.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")